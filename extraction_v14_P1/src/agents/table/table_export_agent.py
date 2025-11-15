#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Table Export Agent - Professional CSV/Excel Export with Notes and Diagrams

This agent handles the final export of extracted table data to multiple formats:
- CSV with notes sections
- Excel with embedded diagrams and notes
- JSON for machine consumption

Design Rationale:
-----------------
- **Format-Specific Logic**: Each format has unique requirements
  - CSV: Append notes as text section below table
  - Excel: Embed diagrams in cells, format notes section
  - JSON: Include all metadata for complete reconstruction

- **Professional Quality**: Excel output matches hand-crafted spreadsheets
  - Proper borders, fonts, alignment
  - Embedded images at appropriate scale
  - Notes section clearly separated

- **Single Responsibility**: Only exports, doesn't extract
  - Takes ExtractedObject as input
  - Produces formatted output files
  - No extraction logic mixed in

Alternatives Considered:
------------------------
1. Export during extraction: Couples extraction and formatting
2. Single format output: Limits user flexibility
3. No diagram embedding: Loses critical visual information

Usage:
------
>>> from table_export_agent import TableExportAgent
>>> exporter = TableExportAgent(output_dir="results/tables")
>>> exporter.export_to_csv(extracted_object)
>>> exporter.export_to_excel([extracted_object1, extracted_object2])
"""

import sys
import os
from pathlib import Path
from typing import List, Dict, Any, Optional
import pandas as pd

# MANDATORY UTF-8 SETUP
if sys.platform == 'win32':
    import io
    if not hasattr(sys.stdout, '_wrapped_utf8'):
        try:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
            sys.stdout._wrapped_utf8 = True
        except (AttributeError, ValueError):
            os.system('chcp 65001')

# Import using proper v14 package structure (no sys.path manipulation)
from common.src.base.base_extraction_agent import ExtractedObject

# TODO: excel_utils will be migrated in future phase
# For now, include minimal validation function directly
def validate_sheet_name(name: str) -> str:
    """
    Validate Excel sheet name to comply with Excel specifications.
    Simple version - full excel_utils will be migrated in future phase.

    Excel requirements:
    - Maximum 31 characters
    - Cannot contain: [ ] : * ? / \\
    """
    import re
    if not name or len(name.strip()) == 0:
        return "Sheet1"

    # Remove invalid characters
    cleaned = re.sub(r'[\[\]:*?/\\]', '_', name)
    cleaned = cleaned.strip().strip("'")

    # Truncate if needed
    if len(cleaned) > 31:
        cleaned = cleaned[:31]

    return cleaned


class TableExportAgent:
    """
    Professional table export to CSV/Excel with notes and diagrams.

    This is NOT a BaseExtractionAgent - it's an export agent that produces
    files from ExtractedObjects.
    """

    def __init__(self, output_dir: Path):
        """
        Initialize export agent.

        Args:
            output_dir: Base directory for output files
        """
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Create subdirectories
        self.csv_dir = self.output_dir / "csv"
        self.excel_dir = self.output_dir / "excel"
        self.csv_dir.mkdir(exist_ok=True)
        self.excel_dir.mkdir(exist_ok=True)

    def export_to_csv(self, obj: ExtractedObject, include_notes: bool = True) -> Path:
        """
        Export table to CSV with optional notes section.

        Args:
            obj: Extracted table object
            include_notes: Whether to append notes section

        Returns:
            Path to created CSV file
        """
        # Get table data
        if 'structured_data' not in obj.content:
            raise ValueError(f"Object {obj.id} missing structured_data")

        structured = obj.content['structured_data']
        headers = structured.get('headers', [])
        rows = structured.get('rows', [])

        # Create DataFrame with mismatch handling
        try:
            df = pd.DataFrame(rows, columns=headers)
        except (ValueError, AssertionError) as e:
            # Handle column/row mismatch - adjust headers or rows
            if rows:
                max_cols = max(len(row) for row in rows)
                if len(headers) != max_cols:
                    # Adjust headers to match data
                    if len(headers) < max_cols:
                        headers = headers + [f'Column {i+1}' for i in range(len(headers), max_cols)]
                    else:
                        headers = headers[:max_cols]
                # Pad short rows
                rows = [row + [''] * (max_cols - len(row)) for row in rows]
            df = pd.DataFrame(rows, columns=headers)

        # Remove "Notes:" marker rows from data
        note_markers = ['notes:', 'note:', 'notes', 'note']
        if len(df.columns) > 0 and len(df) > 0:
            try:
                first_col = df.columns[0]
                # Use iloc to get Series directly (handles duplicate column names)
                mask = df.iloc[:, 0].fillna('').astype(str).str.lower().str.strip().isin(note_markers)
                df = df[~mask]
            except Exception as e:
                # If note filtering fails, continue without filtering
                pass

        # Save CSV
        csv_path = self.csv_dir / f"{obj.id}.csv"
        df.to_csv(csv_path, index=False, encoding='utf-8')

        # Append notes if present
        if include_notes:
            notes = obj.content.get('notes', '')
            if notes and len(notes.strip()) > 0:
                with open(csv_path, 'a', encoding='utf-8') as f:
                    f.write('\n\nNOTES:\n')
                    f.write(notes + '\n')

        return csv_path

    def export_to_excel(
        self,
        objects: List[ExtractedObject],
        excel_filename: str = "tables.xlsx",
        one_sheet_per_table: bool = True
    ) -> Path:
        """
        Export tables to Excel with embedded diagrams and formatted notes.

        Args:
            objects: List of extracted table objects
            excel_filename: Output Excel filename
            one_sheet_per_table: True for separate sheets, False for single sheet

        Returns:
            Path to created Excel file
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.utils.dataframe import dataframe_to_rows

        excel_path = self.excel_dir / excel_filename
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        for obj in objects:
            if obj.type != 'table':
                continue

            # Get table data
            structured = obj.content.get('structured_data', {})
            headers = structured.get('headers', [])
            rows = structured.get('rows', [])

            # Create DataFrame with mismatch handling
            try:
                df = pd.DataFrame(rows, columns=headers)
            except (ValueError, AssertionError) as e:
                # Handle column/row mismatch
                if rows:
                    max_cols = max(len(row) for row in rows)
                    if len(headers) != max_cols:
                        if len(headers) < max_cols:
                            headers = headers + [f'Column {i+1}' for i in range(len(headers), max_cols)]
                        else:
                            headers = headers[:max_cols]
                    rows = [row + [''] * (max_cols - len(row)) for row in rows]
                df = pd.DataFrame(rows, columns=headers)

            # Remove "Notes:" marker rows
            note_markers = ['notes:', 'note:', 'notes', 'note']
            if len(df.columns) > 0 and len(df) > 0:
                try:
                    # Use iloc to get Series directly (handles duplicate column names)
                    mask = df.iloc[:, 0].fillna('').astype(str).str.lower().str.strip().isin(note_markers)
                    df = df[~mask]
                except Exception as e:
                    # If note filtering fails, continue without filtering
                    pass

            # Create worksheet with validated name
            sheet_name = validate_sheet_name(obj.id)  # Intelligent shortening if needed
            ws = wb.create_sheet(title=sheet_name)

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)

                    # Header formatting
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Borders
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )

            # Embed diagrams (images within tables)
            diagrams = obj.content.get('diagrams', [])
            diagram_row = len(df) + 3  # Start after table data with blank row

            if diagrams:
                for diag_idx, diag in enumerate(diagrams):
                    try:
                        img_path = Path(diag.get('image_path', ''))
                        if img_path.exists():
                            img = XLImage(str(img_path))

                            # Scale image to fit Excel (default 0.7 scale, ~70% of original)
                            scale = diag.get('scale', 0.7)
                            if hasattr(img, 'width') and hasattr(img, 'height'):
                                img.width = int(img.width * scale)
                                img.height = int(img.height * scale)

                            # Place image starting at column A of diagram_row
                            ws.add_image(img, f'A{diagram_row}')

                            # Reserve rows for this image (approximate height in Excel rows)
                            # Excel row height is ~15 pixels, so image height / 15 = rows needed
                            rows_needed = max(int(img.height / 15) + 1, 5)  # Minimum 5 rows
                            diagram_row += rows_needed + 1  # Add spacing between images

                            print(f"  ✅ Embedded diagram {diag_idx + 1} from {img_path.name}")
                    except Exception as e:
                        print(f"  ⚠️  Failed to embed diagram {diag_idx + 1}: {e}")
                        continue

            # Add notes section (after diagrams if present)
            notes = obj.content.get('notes', '')
            if notes and len(notes.strip()) > 0:
                notes_start_row = diagram_row if diagrams else len(df) + 3

                # "NOTES:" header
                notes_cell = ws.cell(row=notes_start_row, column=1, value="NOTES:")
                notes_cell.font = Font(bold=True, size=12)

                # Notes content
                notes_lines = notes.split('\n')
                for i, line in enumerate(notes_lines):
                    ws.cell(row=notes_start_row + 1 + i, column=1, value=line)

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save Excel file
        wb.save(excel_path)

        return excel_path

    def export_all(self, objects: List[ExtractedObject]) -> Dict[str, List[Path]]:
        """
        Export tables to all formats.

        Args:
            objects: List of extracted table objects

        Returns:
            Dictionary of format -> list of output paths
        """
        results = {
            'csv': [],
            'excel': []
        }

        # Export individual CSVs
        for obj in objects:
            if obj.type == 'table':
                csv_path = self.export_to_csv(obj)
                results['csv'].append(csv_path)

        # Export combined Excel
        table_objects = [obj for obj in objects if obj.type == 'table']
        if table_objects:
            excel_path = self.export_to_excel(table_objects, excel_filename="all_tables.xlsx")
            results['excel'].append(excel_path)

        return results


if __name__ == "__main__":
    # Simple test with mock data
    print("TableExportAgent standalone testing not implemented.")
    print("Use the orchestrator pipeline for full testing.")
