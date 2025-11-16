# -*- coding: utf-8 -*-
"""
Table Layout Agent - Post-processing for professional table formatting

Handles:
1. Header & unit extraction from table images
2. Row separation and alignment
3. Image enhancement and sizing
4. Professional formatting (borders, fonts, alignment)
"""

import sys
sys.stdout.reconfigure(encoding='utf-8')

import cv2
import numpy as np
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import json
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import easyocr


class TableLayoutConfig:
    """Configuration for table layout formatting."""

    # Image settings
    image_scale_factor = 0.7        # Size of embedded images (0.7 = 70% of original)
    image_border_width = 2          # Border thickness in pixels
    min_row_height_for_image = 120  # Minimum row height in points

    # Header settings
    header_font_size = 11
    header_bold = True
    include_units_in_headers = True

    # Cell formatting
    add_cell_borders = True
    border_style = 'thin'           # 'thin', 'medium', 'thick'
    align_numbers_right = True

    # Column width
    auto_adjust_column_width = True
    min_column_width = 12
    max_column_width = 50


class HeaderExtractor:
    """Extract column headers and units from table crop images."""

    def __init__(self):
        self.reader = easyocr.Reader(['en'], gpu=False)
        self.unit_patterns = [
            r'(Btu/h\s*ft\s*[F°]?)',
            r'(W/m\s*[C°K]?)',
            r'(lb/ft[²³]?)',
            r'(kg/m[²³]?)',
            r'(Btu/h\s*ft[²³]?\s*[F°]?)',
            r'(W/m[²³]?\s*[C°K]?)',
            r'([°]?[FCK])',
        ]

    def extract_headers_from_image(self, table_crop_path: Path) -> List[Dict[str, str]]:
        """
        Extract column headers and units from table image.

        Args:
            table_crop_path: Path to table crop image

        Returns:
            List of dicts with 'name' and 'units' keys
        """
        print(f"  Extracting headers from: {table_crop_path.name}")

        # Load image
        img = cv2.imread(str(table_crop_path))
        if img is None:
            print(f"  ⚠️  Failed to load image")
            return []

        height, width = img.shape[:2]

        # Extract header region (top 20% of image)
        header_height = int(height * 0.20)
        header_img = img[0:header_height, :]

        # OCR the header region
        ocr_results = self.reader.readtext(header_img)

        print(f"  OCR found {len(ocr_results)} text regions in header")

        # Extract column headers by grouping text by x-coordinate
        headers = self._group_headers_by_column(ocr_results, width)

        return headers

    def _group_headers_by_column(self, ocr_results: List, image_width: int) -> List[Dict[str, str]]:
        """Group OCR results into columns."""

        if not ocr_results:
            return []

        # Sort by x-coordinate
        sorted_results = sorted(ocr_results, key=lambda x: x[0][0][0])

        # Simple strategy: assume 3-4 columns, divide image width
        # Better: cluster by x-coordinate gaps

        columns = []
        current_column_texts = []
        last_x = -1000

        for bbox, text, conf in sorted_results:
            x_center = (bbox[0][0] + bbox[2][0]) / 2

            # If gap > 50 pixels, start new column
            if last_x > 0 and (x_center - last_x) > 50:
                if current_column_texts:
                    header = self._parse_column_header(current_column_texts)
                    columns.append(header)
                current_column_texts = []

            current_column_texts.append(text)
            last_x = x_center

        # Add last column
        if current_column_texts:
            header = self._parse_column_header(current_column_texts)
            columns.append(header)

        return columns

    def _parse_column_header(self, texts: List[str]) -> Dict[str, str]:
        """Parse column header to extract name and units."""

        full_text = ' '.join(texts)

        # Try to extract units using patterns
        units = None
        for pattern in self.unit_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                units = match.group(1)
                # Remove units from name
                full_text = re.sub(pattern, '', full_text, flags=re.IGNORECASE).strip()
                break

        # Clean up the name
        name = full_text.strip()
        name = re.sub(r'\s+', ' ', name)  # Normalize whitespace

        return {
            'name': name if name else 'Column',
            'units': units
        }

    def update_excel_headers(self, excel_path: Path, headers: List[Dict[str, str]]):
        """Update Excel file with extracted headers."""

        print(f"  Updating Excel headers...")

        wb = load_workbook(excel_path)
        ws = wb.active

        # Find header row (typically row 3 after title)
        header_row = 3

        # Update column headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)

            # Build header text
            if header['units']:
                header_text = f"{header['name']}, {header['units']}"
            else:
                header_text = header['name']

            cell.value = header_text

            # Format header
            cell.font = Font(bold=True, size=TableLayoutConfig.header_font_size)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            print(f"    Column {col_idx}: {header_text}")

        wb.save(excel_path)
        print(f"  ✅ Headers updated")


class RowAlignmentFixer:
    """Fix row alignment issues like concatenated values."""

    def detect_concatenation(self, df: pd.DataFrame) -> bool:
        """Check if rows have concatenated values."""

        # Look for semicolons in cells (common separator in concatenated data)
        for col in df.columns:
            for val in df[col]:
                if isinstance(val, str) and ';' in val and val.count(';') > 2:
                    return True

        return False

    def split_concatenated_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Split concatenated rows into separate rows."""

        print(f"  Detecting row concatenation...")

        # Find the concatenated row (usually first data row)
        for idx, row in df.iterrows():
            # Check if first column has multiple semicolons
            first_val = row.iloc[0] if len(row) > 0 else None

            if isinstance(first_val, str) and first_val.count(';') > 1:
                print(f"  Found concatenated row at index {idx}")
                return self._split_row(row, df)

        return df

    def _split_row(self, row: pd.Series, original_df: pd.DataFrame) -> pd.DataFrame:
        """Split a single concatenated row."""

        # Strategy: Split by semicolons in first column, then align values
        first_col = str(row.iloc[0])
        conditions = [c.strip() for c in first_col.split(';') if c.strip()]

        print(f"  Split into {len(conditions)} rows")

        # For Table 2, we know the pattern: condition ; h (Btu) ; h (W)
        # Extract values from other columns
        col2_val = str(row.iloc[1]) if len(row) > 1 else ""
        col3_val = str(row.iloc[2]) if len(row) > 2 else ""

        # Split values by whitespace (values are ranges like "2 to 8 5 to 50")
        col2_values = self._extract_ranges(col2_val)
        col3_values = self._extract_ranges(col3_val)

        # Create new rows
        new_rows = []
        for i, condition in enumerate(conditions):
            new_row = {
                original_df.columns[0]: condition,
                original_df.columns[1]: col2_values[i] if i < len(col2_values) else "",
                original_df.columns[2]: col3_values[i] if i < len(col3_values) else ""
            }
            new_rows.append(new_row)

        # Create new DataFrame
        new_df = pd.DataFrame(new_rows)

        # Keep header rows from original
        header_rows = original_df.iloc[:original_df.index[original_df.iloc[:, 0] == row.iloc[0]][0]]

        # Combine
        result = pd.concat([header_rows, new_df], ignore_index=True)

        return result

    def _extract_ranges(self, text: str) -> List[str]:
        """Extract value ranges from concatenated text."""

        # Pattern: "2 to 8 5 to 50 1000 to 3000"
        # or: "6 to 30 30 to 300 1800 to 4800"

        pattern = r'(\d+\s+to\s+\d+)'
        ranges = re.findall(pattern, text)

        return ranges


class ImageEnhancer:
    """Enhance embedded images in Excel files."""

    def __init__(self, config: TableLayoutConfig = None):
        self.config = config or TableLayoutConfig()

    def enhance_images_in_excel(self, excel_path: Path, diagrams_dir: Path, table_num: str):
        """Re-embed images with better sizing and formatting."""

        print(f"  Enhancing images in Table {table_num}...")

        # Load diagram metadata
        metadata_file = diagrams_dir / 'diagram_positions.json'
        if not metadata_file.exists():
            print(f"  No diagrams for Table {table_num}")
            return

        with open(metadata_file, 'r', encoding='utf-8') as f:
            metadata = json.load(f)

        diagrams = metadata.get('diagrams', [])

        if not diagrams:
            print(f"  No diagrams found")
            return

        # Load Excel
        wb = load_workbook(excel_path)
        ws = wb.active

        # Remove existing images
        ws._images = []

        # Re-embed with better sizing
        for diagram in diagrams:
            diagram_filename = diagram.get('filename', '')
            diagram_path = diagrams_dir / diagram_filename

            if not diagram_path.exists():
                continue

            # Create enhanced image
            img = self._create_enhanced_image(diagram_path)

            # Determine position
            row_num, col_num = self._get_image_position(table_num, diagram.get('row', 0))

            # Anchor image
            cell_ref = f"{get_column_letter(col_num)}{row_num}"
            img.anchor = cell_ref

            # Add to worksheet
            ws.add_image(img)

            # Adjust row height
            ws.row_dimensions[row_num].height = max(
                ws.row_dimensions[row_num].height or 0,
                img.height * 0.75 + 10  # Convert pixels to points + padding
            )

            print(f"    Enhanced {diagram_filename} at {cell_ref} (scale {self.config.image_scale_factor})")

        wb.save(excel_path)
        print(f"  ✅ Images enhanced")

    def _create_enhanced_image(self, image_path: Path) -> XLImage:
        """Create enhanced image with borders and proper sizing."""

        # For now, just use openpyxl Image with scale
        # Future: Add border using PIL
        img = XLImage(str(image_path))

        # Apply scale
        img.width = int(img.width * self.config.image_scale_factor)
        img.height = int(img.height * self.config.image_scale_factor)

        return img

    def _get_image_position(self, table_num: str, diagram_row: int) -> Tuple[int, int]:
        """Get Excel row and column for diagram placement."""

        data_start_row = 4  # After title and headers

        if table_num == "4":
            # Table 4: Place diagrams side-by-side below table
            last_row = 10  # Approximate
            row = last_row + 2
            col = 1 + diagram_row

        elif table_num == "5":
            # Table 5: Circuit diagrams in column B (middle)
            row = data_start_row + diagram_row
            col = 2

        elif table_num == "6":
            # Table 6: Geometric diagrams in column A (left)
            row = data_start_row + diagram_row
            col = 1

        else:
            row = data_start_row
            col = 1

        return row, col


class FormattingEngine:
    """Apply professional formatting to Excel tables."""

    def __init__(self, config: TableLayoutConfig = None):
        self.config = config or TableLayoutConfig()

    def apply_formatting(self, excel_path: Path):
        """Apply professional formatting to Excel file."""

        print(f"  Applying professional formatting...")

        wb = load_workbook(excel_path)
        ws = wb.active

        # Apply borders to all cells
        if self.config.add_cell_borders:
            self._apply_borders(ws)

        # Adjust column widths
        if self.config.auto_adjust_column_width:
            self._adjust_column_widths(ws)

        # Align cell content
        self._align_cells(ws)

        wb.save(excel_path)
        print(f"  ✅ Formatting applied")

    def _apply_borders(self, ws):
        """Apply borders to all cells with data."""

        border_style = self.config.border_style
        border = Border(
            left=Side(style=border_style),
            right=Side(style=border_style),
            top=Side(style=border_style),
            bottom=Side(style=border_style)
        )

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    cell.border = border

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""

        for col_idx in range(1, ws.max_column + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)

            # Set width with constraints
            adjusted_width = min(
                max(max_length + 2, self.config.min_column_width),
                self.config.max_column_width
            )

            ws.column_dimensions[column_letter].width = adjusted_width

    def _align_cells(self, ws):
        """Align cell content based on type."""

        for row in ws.iter_rows(min_row=4):  # Skip title and headers
            for cell in row:
                if cell.value:
                    # Check if number
                    try:
                        float(str(cell.value).replace(',', '').replace('to', ''))
                        # Number - right align
                        if self.config.align_numbers_right:
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    except ValueError:
                        # Text - left align
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)


class TableLayoutAgent:
    """Main agent orchestrating all table layout improvements."""

    def __init__(self, config: TableLayoutConfig = None):
        self.config = config or TableLayoutConfig()
        self.header_extractor = HeaderExtractor()
        self.row_fixer = RowAlignmentFixer()
        self.image_enhancer = ImageEnhancer(self.config)
        self.formatter = FormattingEngine(self.config)

    def process_table(
        self,
        excel_path: Path,
        table_crop_path: Path,
        table_num: str,
        diagrams_dir: Optional[Path] = None,
        fix_headers: bool = True,
        fix_rows: bool = True,
        enhance_images: bool = True,
        apply_formatting: bool = True
    ):
        """
        Process a single table through all layout improvements.

        Args:
            excel_path: Path to Excel file
            table_crop_path: Path to table crop image
            table_num: Table number
            diagrams_dir: Path to diagrams directory (if applicable)
            fix_headers: Whether to extract and update headers
            fix_rows: Whether to fix row alignment
            enhance_images: Whether to enhance embedded images
            apply_formatting: Whether to apply professional formatting
        """

        print(f"\n{'='*70}")
        print(f"Processing Table {table_num}")
        print(f"{'='*70}")
        print(f"  Excel: {excel_path.name}")
        print(f"  Crop: {table_crop_path.name}")

        # Step 1: Extract and update headers
        if fix_headers:
            headers = self.header_extractor.extract_headers_from_image(table_crop_path)
            if headers:
                self.header_extractor.update_excel_headers(excel_path, headers)

        # Step 2: Fix row alignment
        if fix_rows:
            df = pd.read_excel(excel_path)
            if self.row_fixer.detect_concatenation(df):
                fixed_df = self.row_fixer.split_concatenated_rows(df)
                # Save back to Excel
                fixed_df.to_excel(excel_path, index=False)
                print(f"  ✅ Row alignment fixed")

        # Step 3: Enhance images
        if enhance_images and diagrams_dir and diagrams_dir.exists():
            self.image_enhancer.enhance_images_in_excel(excel_path, diagrams_dir, table_num)

        # Step 4: Apply formatting
        if apply_formatting:
            self.formatter.apply_formatting(excel_path)

        print(f"✅ Table {table_num} processing complete")


def main():
    """Process all tables with layout improvements."""

    print("="*70)
    print("TABLE LAYOUT AGENT - ITERATIVE IMPROVEMENTS")
    print("="*70)

    config = TableLayoutConfig()
    agent = TableLayoutAgent(config)

    tables_dir = Path('table_extractions_v3')
    crops_dir = Path('extractions/doclayout_tables')
    diagrams_base = Path('table_extractions_v2/diagrams')

    # Priority 1: Tables with header/unit issues
    print("\n" + "="*70)
    print("PRIORITY 1: HEADER & UNIT EXTRACTION")
    print("="*70)

    # Table 1
    agent.process_table(
        excel_path=tables_dir / 'table_1.xlsx',
        table_crop_path=crops_dir / 'table_1_page2.png',
        table_num='1',
        fix_headers=True,
        fix_rows=False,
        enhance_images=False,
        apply_formatting=True
    )

    # Table 3
    agent.process_table(
        excel_path=tables_dir / 'table_3.xlsx',
        table_crop_path=crops_dir / 'table_3_page5.png',
        table_num='3',
        fix_headers=True,
        fix_rows=False,
        enhance_images=False,
        apply_formatting=True
    )

    # Priority 2: Row alignment issues
    print("\n" + "="*70)
    print("PRIORITY 2: ROW ALIGNMENT")
    print("="*70)

    # Table 2
    agent.process_table(
        excel_path=tables_dir / 'table_2.xlsx',
        table_crop_path=crops_dir / 'table_2_page4.png',
        table_num='2',
        fix_headers=True,
        fix_rows=True,
        enhance_images=False,
        apply_formatting=True
    )

    # Priority 3: Image enhancement
    print("\n" + "="*70)
    print("PRIORITY 3: IMAGE ENHANCEMENT")
    print("="*70)

    # Table 4
    agent.process_table(
        excel_path=tables_dir / 'table_4.xlsx',
        table_crop_path=crops_dir / 'table_unnumbered_1_page7.png',
        table_num='4',
        diagrams_dir=diagrams_base / 'table_04',
        fix_headers=True,
        fix_rows=False,
        enhance_images=True,
        apply_formatting=True
    )

    # Table 5
    agent.process_table(
        excel_path=tables_dir / 'table_5.xlsx',
        table_crop_path=crops_dir / 'table_5_page9.png',
        table_num='5',
        diagrams_dir=diagrams_base / 'table_05',
        fix_headers=True,
        fix_rows=False,
        enhance_images=True,
        apply_formatting=True
    )

    # Table 6
    agent.process_table(
        excel_path=tables_dir / 'table_6.xlsx',
        table_crop_path=crops_dir / 'table_6_page9.png',
        table_num='6',
        diagrams_dir=diagrams_base / 'table_06',
        fix_headers=True,
        fix_rows=False,
        enhance_images=True,
        apply_formatting=True
    )

    print("\n" + "="*70)
    print("TABLE LAYOUT IMPROVEMENTS COMPLETE")
    print("="*70)
    print(f"\nUpdated tables in: {tables_dir.absolute()}")
    print("\nImprovements applied:")
    print("  ✅ Tables 1, 3: Headers and units extracted")
    print("  ✅ Table 2: Row alignment fixed")
    print("  ✅ Tables 4, 5, 6: Images enhanced (scale 0.7)")
    print("  ✅ All tables: Professional formatting applied")
    print("\nNext: Review Excel files and provide feedback")


if __name__ == '__main__':
    main()
